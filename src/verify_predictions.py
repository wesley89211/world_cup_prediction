"""
verify_predictions.py  （滾動更新版 v3）
=========================================
每場比賽結束後自動更新模型狀態，讓後續預測更準確。

滾動更新的內容：
  1. ELO 分數  → 每場 WC 比賽後即時更新（K=64，反映大賽重要性）
  2. Form 分數 → 世界盃場次加入近期狀態分計算
  3. 積分情境  → 偵測晉級/淘汰情況，調整 Poisson λ
               已確定晉級 → 進球預測 × 0.85（保守踢）
               必須贏才能晉級 → 進球預測 × 1.15（拚命踢）

═══════════════════════════════════════
  每日標準流程
═══════════════════════════════════════

  【第一步：只需執行一次】
  python src/verify_predictions.py --predict

  【早上：查看今天有哪些比賽】
  python src/verify_predictions.py --today

  【賽後：輸入比分，自動更新模型狀態並重算後續預測】
  python src/verify_predictions.py --verify

  【賠率 vs 模型對比（需設定 BSD_API_KEY）】
  python src/verify_predictions.py --odds

  【查看累計準確率】
  python src/verify_predictions.py --stats

  搭配 --date YYYY-MM-DD 指定台灣日期
"""

import pandas as pd
import numpy as np
import pickle
import argparse
from pathlib import Path
from datetime import datetime, timedelta, timezone
from scipy.stats import poisson as scipy_poisson
try:
    import requests as _requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ── 路徑設定 ─────────────────────────────────────────────────────
XGB_PATH      = r"data/models/xgb_v2.pkl"
POISSON_PATH  = r"data/models/advanced_poisson_v2.pkl"
ENSEMBLE_PATH = r"data/models/ensemble_v2.pkl"
FEATURES_PATH = r"data/processed/wc26_final_features.csv"
ELO_PATH      = r"data/processed/wc26_elo_v2.csv"
H2H_PATH      = r"data/processed/wc26_h2h.csv"
PRED_LOG      = r"data/processed/wc2026_predictions.csv"
LIVE_STATE    = r"data/processed/wc2026_live_state.csv"   # 滾動狀態存檔
ODDS_EXCEL    = r"data/reports/wc2026_odds_comparison.xlsx"  # 賠率比較輸出

# BSD API 設定
BSD_API_KEY       = ""          # ← 填入你的 BSD API key
BSD_BASE          = "https://sports.bzzoiro.com"
BSD_WC_LEAGUE_ID  = 27          # World Cup 2026
BSD_VALUE_BET_THR = 0.05        # Value Bet 閾值（模型 - 市場 > 5%）
BSD_NAME_MAP = {                # BSD 隊名 → 我們的隊名
    "Czechia":       "Czech Republic",
    "Türkiye":       "Turkey",
    "Côte d'Ivoire": "Ivory Coast",
    "Cabo Verde":    "Cape Verde",
    "USA":           "United States",
}
# ─────────────────────────────────────────────────────────────────

WC_GOAL_SCALE = 1.35   # 世界盃進球校正係數
WC_ELO_K      = 64     # 世界盃 ELO 更新幅度（比一般賽事 K=32 更高）
FORM_DECAY    = 0.85   # 近期 form 衰減係數

# ── 賽程表（含美東→台灣時間轉換）───────────────────────────────
_FIXTURES_RAW = [
    ("2026-06-11","15:00","Mexico",        "South Africa",          "A"),
    ("2026-06-11","22:00","South Korea",   "Czech Republic",        "A"),
    ("2026-06-18","12:00","Czech Republic","South Africa",          "A"),
    ("2026-06-18","21:00","Mexico",        "South Korea",           "A"),
    ("2026-06-24","21:00","Czech Republic","Mexico",                "A"),
    ("2026-06-24","21:00","South Africa",  "South Korea",           "A"),
    ("2026-06-12","15:00","Canada",        "Bosnia and Herzegovina","B"),
    ("2026-06-13","15:00","Qatar",         "Switzerland",           "B"),
    ("2026-06-18","15:00","Switzerland",   "Bosnia and Herzegovina","B"),
    ("2026-06-18","18:00","Canada",        "Qatar",                 "B"),
    ("2026-06-24","15:00","Switzerland",   "Canada",                "B"),
    ("2026-06-24","15:00","Bosnia and Herzegovina","Qatar",         "B"),
    ("2026-06-13","18:00","Brazil",        "Morocco",               "C"),
    ("2026-06-13","21:00","Haiti",         "Scotland",              "C"),
    ("2026-06-19","18:00","Scotland",      "Morocco",               "C"),
    ("2026-06-19","21:00","Brazil",        "Haiti",                 "C"),
    ("2026-06-24","18:00","Scotland",      "Brazil",                "C"),
    ("2026-06-24","18:00","Morocco",       "Haiti",                 "C"),
    ("2026-06-12","21:00","United States", "Paraguay",              "D"),
    ("2026-06-14","00:00","Australia",     "Turkey",                "D"),
    ("2026-06-19","15:00","United States", "Australia",             "D"),
    ("2026-06-19","18:00","Paraguay",      "Turkey",                "D"),
    ("2026-06-24","21:00","Turkey",        "United States",         "D"),
    ("2026-06-24","21:00","Paraguay",      "Australia",             "D"),
    ("2026-06-14","13:00","Germany",       "Curaçao",               "E"),
    ("2026-06-15","01:00","Ivory Coast",   "Ecuador",               "E"),
    ("2026-06-20","15:00","Germany",       "Ivory Coast",           "E"),
    ("2026-06-20","18:00","Ecuador",       "Curaçao",               "E"),
    ("2026-06-25","18:00","Curaçao",       "Ecuador",               "E"),
    ("2026-06-25","18:00","Ivory Coast",   "Germany",               "E"),
    ("2026-06-14","16:00","Netherlands",   "Japan",                 "F"),
    ("2026-06-15","04:00","Sweden",        "Tunisia",               "F"),
    ("2026-06-20","12:00","Netherlands",   "Sweden",                "F"),
    ("2026-06-20","21:00","Japan",         "Tunisia",               "F"),
    ("2026-06-25","15:00","Tunisia",       "Netherlands",           "F"),
    ("2026-06-25","21:00","Japan",         "Sweden",                "F"),
    ("2026-06-15","15:00","Belgium",       "Egypt",                 "G"),
    ("2026-06-15","21:00","Iran",          "New Zealand",           "G"),
    ("2026-06-21","15:00","Belgium",       "Iran",                  "G"),
    ("2026-06-21","18:00","New Zealand",   "Egypt",                 "G"),
    ("2026-06-26","23:00","Egypt",         "Iran",                  "G"),
    ("2026-06-26","23:00","New Zealand",   "Belgium",               "G"),
    ("2026-06-15","12:00","Spain",         "Cape Verde",            "H"),
    ("2026-06-15","18:00","Saudi Arabia",  "Uruguay",               "H"),
    ("2026-06-21","12:00","Spain",         "Saudi Arabia",          "H"),
    ("2026-06-21","21:00","Uruguay",       "Cape Verde",            "H"),
    ("2026-06-26","15:00","Cape Verde",    "Saudi Arabia",          "H"),
    ("2026-06-26","15:00","Uruguay",       "Spain",                 "H"),
    ("2026-06-16","15:00","France",        "Senegal",               "I"),
    ("2026-06-16","18:00","Iraq",          "Norway",                "I"),
    ("2026-06-22","15:00","France",        "Iraq",                  "I"),
    ("2026-06-22","18:00","Norway",        "Senegal",               "I"),
    ("2026-06-27","15:00","Senegal",       "Iraq",                  "I"),
    ("2026-06-27","15:00","Norway",        "France",                "I"),
    ("2026-06-16","21:00","Argentina",     "Algeria",               "J"),
    ("2026-06-16","00:00","Austria",       "Jordan",                "J"),
    ("2026-06-22","21:00","Argentina",     "Austria",               "J"),
    ("2026-06-22","12:00","Jordan",        "Algeria",               "J"),
    ("2026-06-27","21:00","Algeria",       "Austria",               "J"),
    ("2026-06-27","21:00","Jordan",        "Argentina",             "J"),
    ("2026-06-17","13:00","Portugal",      "DR Congo",              "K"),
    ("2026-06-17","22:00","Uzbekistan",    "Colombia",              "K"),
    ("2026-06-23","13:00","Portugal",      "Uzbekistan",            "K"),
    ("2026-06-23","22:00","Colombia",      "DR Congo",              "K"),
    ("2026-06-28","15:00","DR Congo",      "Uzbekistan",            "K"),
    ("2026-06-28","15:00","Colombia",      "Portugal",              "K"),
    ("2026-06-17","16:00","England",       "Croatia",               "L"),
    ("2026-06-17","19:00","Ghana",         "Panama",                "L"),
    ("2026-06-23","16:00","England",       "Ghana",                 "L"),
    ("2026-06-23","19:00","Panama",        "Croatia",               "L"),
    ("2026-06-28","18:00","Croatia",       "Ghana",                 "L"),
    ("2026-06-28","18:00","Panama",        "England",               "L"),
]

def _build_fixtures():
    fixtures = []
    for et_date, et_time, home, away, group in _FIXTURES_RAW:
        et_dt = datetime.strptime(f"{et_date} {et_time}", "%Y-%m-%d %H:%M")
        tw_dt = et_dt + timedelta(hours=12)
        fixtures.append({
            "et_date": et_date, "tw_date": tw_dt.strftime("%Y-%m-%d"),
            "tw_time": tw_dt.strftime("%H:%M"), "tw_datetime": tw_dt,
            "home": home, "away": away, "group": group, "neutral": True,
        })
    return fixtures

WC2026_FIXTURES = _build_fixtures()


# ════════════════════════════════════════════════════════════════
# 滾動狀態管理
# ════════════════════════════════════════════════════════════════

def load_live_state():
    """
    載入滾動狀態（ELO、form、積分）
    若不存在則從初始值建立
    """
    state_path = Path(LIVE_STATE)

    # 初始 ELO
    elo_df = pd.read_csv(ELO_PATH)
    init_elo = dict(zip(elo_df["team"], elo_df["elo_weighted"]))

    # 初始 form（從 wc26_final_features 取）
    feat_df = pd.read_csv(FEATURES_PATH)
    init_form = dict(zip(feat_df["team"], feat_df["form_score_official"].fillna(0.5)))

    if not state_path.exists():
        # 初始化所有隊伍的狀態
        all_teams = list(set(
            [f["home"] for f in WC2026_FIXTURES] +
            [f["away"] for f in WC2026_FIXTURES]
        ))
        rows = []
        for team in all_teams:
            grp = next((f["group"] for f in WC2026_FIXTURES if f["home"]==team or f["away"]==team), "?")
            rows.append({
                "team": team, "group": grp,
                "elo": round(init_elo.get(team, 1700), 1),
                "form": round(init_form.get(team, 0.5), 4),
                "pts": 0, "w": 0, "d": 0, "l": 0,
                "gf": 0, "ga": 0, "played": 0,
            })
        state = pd.DataFrame(rows)
        state.to_csv(LIVE_STATE, index=False, encoding="utf-8-sig")
        return state

    return pd.read_csv(LIVE_STATE)


def save_live_state(state: pd.DataFrame):
    state.to_csv(LIVE_STATE, index=False, encoding="utf-8-sig")


def update_live_state(state: pd.DataFrame, home: str, away: str,
                      hg: int, ag: int, group: str) -> pd.DataFrame:
    """
    比賽結束後更新滾動狀態：
    1. ELO（WC K=64）
    2. form score（加入最新結果，指數衰減）
    3. 積分榜（W/D/L/GF/GA）
    """
    state = state.copy()

    def get_row(team):
        mask = state["team"] == team
        if not mask.any():
            state.loc[len(state)] = {"team":team,"group":group,"elo":1700,"form":0.5,"pts":0,"w":0,"d":0,"l":0,"gf":0,"ga":0,"played":0}
        return state[state["team"] == team].index[0]

    hi = get_row(home)
    ai = get_row(away)

    he = float(state.at[hi, "elo"])
    ae = float(state.at[ai, "elo"])

    # ── ELO 更新 ──
    we_h = 1 / (1 + 10**((ae - he) / 400))
    if hg > ag:   wh, wa = 1.0, 0.0
    elif hg == ag: wh, wa = 0.5, 0.5
    else:          wh, wa = 0.0, 1.0

    state.at[hi, "elo"] = round(he + WC_ELO_K * (wh - we_h), 1)
    state.at[ai, "elo"] = round(ae + WC_ELO_K * (wa - (1-we_h)), 1)

    # ── Form 更新（指數衰減，最新一場加進來）──
    def update_form(idx, result_pts):
        old_form = float(state.at[idx, "form"])
        # 新比賽得分：勝=1.0, 平=0.333, 負=0.0
        new_result = result_pts / 3.0
        # 衰減舊 form，加入新結果
        new_form = FORM_DECAY * old_form + (1 - FORM_DECAY) * new_result
        state.at[idx, "form"] = round(new_form, 4)

    update_form(hi, 3 if hg>ag else (1 if hg==ag else 0))
    update_form(ai, 3 if ag>hg else (1 if hg==ag else 0))

    # ── 積分更新 ──
    state.at[hi, "played"] += 1
    state.at[ai, "played"] += 1
    state.at[hi, "gf"] += hg; state.at[hi, "ga"] += ag
    state.at[ai, "gf"] += ag; state.at[ai, "ga"] += hg

    if hg > ag:
        state.at[hi, "pts"] += 3; state.at[hi, "w"] += 1
        state.at[ai, "l"] += 1
    elif hg == ag:
        state.at[hi, "pts"] += 1; state.at[hi, "d"] += 1
        state.at[ai, "pts"] += 1; state.at[ai, "d"] += 1
    else:
        state.at[ai, "pts"] += 3; state.at[ai, "w"] += 1
        state.at[hi, "l"] += 1

    return state


def get_group_standings(state: pd.DataFrame, group: str) -> pd.DataFrame:
    """取得指定組的積分榜"""
    grp = state[state["group"] == group].copy()
    grp["gd"] = grp["gf"] - grp["ga"]
    return grp.sort_values(["pts","gd","gf"], ascending=[False,False,False]).reset_index(drop=True)


def get_motivation_scale(state: pd.DataFrame, team: str, group: str) -> float:
    """
    根據積分情況判斷球隊「拚命程度」
    - 已確定晉級（第1且積分遙遙領先）→ 0.85（保守踢，不傷球員）
    - 必須贏才能晉級 → 1.15（全力以赴）
    - 正常情況 → 1.00
    """
    standings = get_group_standings(state, group)
    if len(standings) < 2:
        return 1.0

    team_row = standings[standings["team"] == team]
    if team_row.empty:
        return 1.0

    rank = team_row.index[0] + 1  # 1-based
    played = int(team_row.iloc[0]["played"])
    pts = int(team_row.iloc[0]["pts"])

    # 3場賽完後才評估（第2輪開始）
    if played == 0:
        return 1.0

    # 最大可得分
    remaining = 3 - played
    max_pts = pts + remaining * 3

    # 第2名的分數
    if len(standings) >= 2:
        second_pts = int(standings.iloc[1]["pts"])
        first_pts  = int(standings.iloc[0]["pts"])
    else:
        return 1.0

    # 已確定晉級：第1名且即使輸完剩下的也能晉級
    if rank == 1 and played >= 2 and pts >= second_pts + 4:
        return 0.85  # 保守踢

    # 必須贏：還有機會但必須全力以赴
    if rank >= 3 and remaining >= 1 and max_pts >= second_pts:
        return 1.15  # 拚命踢

    # 已確定淘汰
    if rank == 4 and max_pts < second_pts:
        return 0.90  # 意興闌珊

    return 1.00


# ════════════════════════════════════════════════════════════════
# 模型載入 & 特徵建立
# ════════════════════════════════════════════════════════════════

def load_models():
    with open(XGB_PATH,      "rb") as f: xgb_data = pickle.load(f)
    with open(POISSON_PATH,  "rb") as f: poi_data = pickle.load(f)
    with open(ENSEMBLE_PATH, "rb") as f: ens_data = pickle.load(f)
    return xgb_data, poi_data, ens_data


def load_features():
    feat = pd.read_csv(FEATURES_PATH).set_index("team")
    h2h_raw = pd.read_csv(H2H_PATH)
    h2h_dict = {}
    for _, r in h2h_raw.iterrows():
        a, b, t = r["team_a"], r["team_b"], r["total_matches"]
        if t == 0: continue
        h2h_dict[(a,b)] = {"h2h_total":t,"h2h_home_winrate":r["team_a_wins"]/t,"h2h_draw_rate":r["draws"]/t}
        h2h_dict[(b,a)] = {"h2h_total":t,"h2h_home_winrate":r["team_b_wins"]/t,"h2h_draw_rate":r["draws"]/t}
    return feat, h2h_dict


def build_features(home, away, feat, state, h2h_dict):
    """用靜態特徵表 + 滾動狀態建立特徵向量"""
    feat_means  = feat.mean()
    ELO_DEFAULT = 1700

    def get_f(team, col):
        if team in feat.index:
            v = feat.at[team, col]
            return v if pd.notna(v) else feat_means[col]
        return feat_means[col]

    # 從滾動狀態取最新 ELO 和 form
    def get_live(team, col, default):
        row = state[state["team"] == team]
        return float(row.iloc[0][col]) if not row.empty else default

    he = get_live(home, "elo",  ELO_DEFAULT)
    ae = get_live(away, "elo",  ELO_DEFAULT)
    hf = get_live(home, "form", 0.5)
    af = get_live(away, "form", 0.5)

    h2h = h2h_dict.get((home, away), {"h2h_total":0,"h2h_home_winrate":0.333,"h2h_draw_rate":0.25})

    row = {
        "home_elo": he, "away_elo": ae, "elo_diff": he - ae,
        "is_neutral": 1, "is_wc_final": 1, "is_friendly": 0,
        "h2h_home_winrate": h2h["h2h_home_winrate"],
        "h2h_draw_rate":    h2h["h2h_draw_rate"],
        "h2h_total":        h2h["h2h_total"],
    }
    for col in feat.columns:
        row[f"home_{col}"] = get_f(home, col)
        row[f"away_{col}"] = get_f(away, col)

    # 用滾動 form 蓋掉靜態 form
    row["home_form_score_official"] = hf
    row["away_form_score_official"] = af
    row["diff_form_score_official"] = hf - af

    diff_targets = ["total_value_eur","game_top_11_ovr","game_fw_speed","game_fw_finishing",
                    "game_mf_passing","game_df_defense","game_df_physic","win_rate_r5","win_rate_wc",
                    "goal_diff_avg_r5","goals_for_avg_r5","goals_against_avg_r5","top_club_ratio","avg_age"]
    for col in diff_targets:
        if f"home_{col}" in row:
            row[f"diff_{col}"] = row[f"home_{col}"] - row[f"away_{col}"]
    row["diff_value_M"] = row.get("diff_total_value_eur", 0) / 1e6
    return row


def poisson_probs(lh, la, max_goals=8):
    p_h = p_d = p_a = 0.0
    for gh in range(max_goals+1):
        for ga in range(max_goals+1):
            p = scipy_poisson.pmf(gh,lh)*scipy_poisson.pmf(ga,la)
            if gh>ga: p_h+=p
            elif gh==ga: p_d+=p
            else: p_a+=p
    t = p_h+p_d+p_a
    return p_h/t, p_d/t, p_a/t


def predict_match(home, away, group, xgb_data, poi_data, ens_data, feat, state, h2h_dict):
    """預測單場（使用滾動狀態中的最新 ELO + form）"""
    row = build_features(home, away, feat, state, h2h_dict)

    # XGBoost
    X_xgb    = np.array([[row.get(f,0) for f in xgb_data["features"]]])
    xgb_prob = xgb_data["model"].predict_proba(X_xgb)[0]

    # Poisson
    X_poi   = np.array([[row.get(f,0) for f in poi_data["poi_cols"]]])
    X_poi_s = poi_data["scaler"].transform(X_poi)
    lh_raw  = max(poi_data["pr_home"].predict(X_poi_s)[0], 0.05)
    la_raw  = max(poi_data["pr_away"].predict(X_poi_s)[0], 0.05)
    ph, pd_, pa = poisson_probs(lh_raw, la_raw)
    poi_prob = np.array([pa, pd_, ph])

    # Ensemble
    w        = ens_data["xgb_weight"]
    ens_prob = w*xgb_prob + (1-w)*poi_prob

    # 積分情境調整（拚命程度）
    mh = get_motivation_scale(state, home, group)
    ma = get_motivation_scale(state, away, group)

    # 比分 Top 3（WC 校正 + 積分情境）
    lh = lh_raw * WC_GOAL_SCALE * mh
    la = la_raw * WC_GOAL_SCALE * ma
    sp = [(gh,ga,round(scipy_poisson.pmf(gh,lh)*scipy_poisson.pmf(ga,la)*100,1))
          for gh in range(11) for ga in range(11)]
    sp.sort(key=lambda x: -x[2])
    t3 = sp[:3]

    pred_result = ("H" if ens_prob[2]>ens_prob[0] and ens_prob[2]>ens_prob[1]
                   else ("D" if ens_prob[1]>=ens_prob[0] and ens_prob[1]>=ens_prob[2] else "A"))

    # 積分情境說明
    motivation_note = ""
    if mh < 1.0: motivation_note = f"  ⚠ {home} 已確定晉級，可能保守踢（進球預測 ×{mh}）"
    if ma < 1.0: motivation_note += f"\n  ⚠ {away} 已確定晉級，可能保守踢（進球預測 ×{ma}）"
    if mh > 1.0: motivation_note = f"  🔥 {home} 必須贏，全力以赴（進球預測 ×{mh}）"
    if ma > 1.0: motivation_note += f"\n  🔥 {away} 必須贏，全力以赴（進球預測 ×{ma}）"

    return {
        "p_home":round(ens_prob[2]*100,1), "p_draw":round(ens_prob[1]*100,1), "p_away":round(ens_prob[0]*100,1),
        "pred_result":pred_result,
        "pred_home_goals":t3[0][0],"pred_away_goals":t3[0][1],"score1_prob":t3[0][2],
        "score2_home":t3[1][0],"score2_away":t3[1][1],"score2_prob":t3[1][2],
        "score3_home":t3[2][0],"score3_away":t3[2][1],"score3_prob":t3[2][2],
        "lambda_home":round(lh,2),"lambda_away":round(la,2),
        "motivation_note": motivation_note,
        "home_elo": round(float(state[state["team"]==home]["elo"].iloc[0]) if not state[state["team"]==home].empty else 1700, 1),
        "away_elo": round(float(state[state["team"]==away]["elo"].iloc[0]) if not state[state["team"]==away].empty else 1700, 1),
    }



# ════════════════════════════════════════════════════════════════
# BSD 賠率整合
# ════════════════════════════════════════════════════════════════

def _fetch_bsd_events() -> list:
    """從 BSD API 拉取 WC 比賽資料（含賠率）"""
    if not HAS_REQUESTS:
        print("  ⚠️  未安裝 requests，跳過賠率拉取")
        return []
    if not BSD_API_KEY:
        print("  ⚠️  BSD_API_KEY 未設定，跳過賠率拉取")
        return []
    headers = {"Authorization": f"Token {BSD_API_KEY}"}
    all_events, url = [], f"{BSD_BASE}/api/events/"
    params = {"league": BSD_WC_LEAGUE_ID, "limit": 100}
    while url:
        try:
            r = _requests.get(url, headers=headers, params=params, timeout=10)
            if r.status_code != 200:
                print(f"  ⚠️  BSD API {r.status_code}")
                break
            data = r.json()
            all_events.extend(data.get("results", []))
            url, params = data.get("next"), {}
        except Exception as e:
            print(f"  ⚠️  BSD 連線失敗：{e}")
            break
    return all_events


def _bsd_implied_prob(odds_h, odds_d, odds_a):
    """decimal 賠率 → 去水份隱含機率（%）"""
    try:
        rh, rd, ra = 1/float(odds_h), 1/float(odds_d), 1/float(odds_a)
        t = rh + rd + ra
        return round(rh/t*100,1), round(rd/t*100,1), round(ra/t*100,1)
    except (TypeError, ZeroDivisionError, ValueError):
        return None, None, None


def _bsd_vig(odds_h, odds_d, odds_a):
    try:
        return round((1/float(odds_h)+1/float(odds_d)+1/float(odds_a)-1)*100, 2)
    except (TypeError, ZeroDivisionError, ValueError):
        return None


def cmd_odds(tw_date_filter=None, export=True):
    """
    拉取 BSD 賠率，與模型預測比對，顯示並可輸出 Excel。
    --odds           今天（台灣時間）的賠率 vs 模型
    --odds --date X  指定台灣日期
    --odds --export  同時輸出 Excel（預設開啟）
    """
    target = tw_date_filter or (datetime.now(timezone.utc)+timedelta(hours=8)).strftime("%Y-%m-%d")

    print(f"\n📡 拉取 BSD 賠率（World Cup 2026）...")
    events = _fetch_bsd_events()
    if not events:
        print("  無法取得賠率資料"); return

    # 轉成 dict，key=(home_mapped, away_mapped)
    bsd_map = {}
    for e in events:
        h = BSD_NAME_MAP.get(e["home_team"], e["home_team"])
        a = BSD_NAME_MAP.get(e["away_team"], e["away_team"])
        bsd_map[(h, a)] = e

    # 讀模型預測
    pred_log = Path(PRED_LOG)
    pred_df  = pd.read_csv(PRED_LOG) if pred_log.exists() else pd.DataFrame()

    # 篩今天的比賽
    fixtures_today = [f for f in WC2026_FIXTURES if f["tw_date"] == target]
    if not fixtures_today:
        print(f"  台灣時間 {target} 沒有分組賽")
        return

    print(f"\n{'='*65}")
    print(f"  賠率 vs 模型  │  台灣時間 {target}")
    print(f"  【市場機率】來自去水份後的 BSD 賠率隱含機率")
    print(f"  【模型機率】來自 XGBoost + Poisson Ensemble")
    print(f"  【Value Bet🔥】模型機率 > 市場機率 ≥ {BSD_VALUE_BET_THR*100:.0f}%")
    print(f"{'='*65}")

    rows_out = []
    for fix in sorted(fixtures_today, key=lambda x: x["tw_time"]):
        home, away = fix["home"], fix["away"]
        e = bsd_map.get((home, away))

        # 賠率
        if e and e.get("odds_home"):
            odds_h, odds_d, odds_a = e["odds_home"], e["odds_draw"], e["odds_away"]
            mkt_h, mkt_d, mkt_a   = _bsd_implied_prob(odds_h, odds_d, odds_a)
            vig                    = _bsd_vig(odds_h, odds_d, odds_a)
            odds_str = f"H={odds_h}  D={odds_d}  A={odds_a}  (水份{vig}%)"
            mkt_str  = f"主場勝 {mkt_h}%  平局 {mkt_d}%  客場勝 {mkt_a}%"
        else:
            odds_h = odds_d = odds_a = None
            mkt_h = mkt_d = mkt_a = vig = None
            odds_str = "（無賠率）"
            mkt_str  = ""

        # 模型預測
        pred_row = None
        if not pred_df.empty:
            m = pred_df[(pred_df["home"]==home) & (pred_df["away"]==away)]
            if not m.empty:
                pred_row = m.iloc[0]

        print(f"\n  {fix['tw_time']}  Group {fix['group']}  {home} vs {away}")

        # 實際比分（已知的話顯示）
        if pred_row is not None and str(pred_row.get("actual_result","")).strip() in ("H","D","A"):
            result_icon = "✅" if pred_row["correct"] else "❌"
            print(f"  實際結果 ▶ {int(pred_row['actual_home_score'])}-{int(pred_row['actual_away_score'])}  {result_icon}")

        if odds_str != "（無賠率）":
            print(f"  ── 賠率  ──  {odds_str}")
            print(f"  ── 市場  ──  {mkt_str}")

        if pred_row is not None:
            p_h = float(pred_row["p_home"])
            p_d = float(pred_row["p_draw"])
            p_a = float(pred_row["p_away"])
            winner = home if pred_row["pred_result"]=="H" else ("平局" if pred_row["pred_result"]=="D" else away)
            print(f"  ── 模型  ──  主場勝 {p_h}%  平局 {p_d}%  客場勝 {p_a}%  → {winner}")
            print(f"  ── 比分  ──  #{1} {int(pred_row['pred_home_goals'])}-{int(pred_row['pred_away_goals'])}({pred_row['score1_prob']}%)"
                  f"  #{2} {int(pred_row['score2_home'])}-{int(pred_row['score2_away'])}({pred_row['score2_prob']}%)"
                  f"  #{3} {int(pred_row['score3_home'])}-{int(pred_row['score3_away'])}({pred_row['score3_prob']}%)")

            # Value Bet 偵測
            vb_hits = []
            if mkt_h:
                for label, mp, mkp in [("主場勝",p_h,mkt_h),("平局",p_d,mkt_d),("客場勝",p_a,mkt_a)]:
                    if mp and mkp and (mp - mkp) >= BSD_VALUE_BET_THR*100:
                        vb_hits.append(f"{label} +{mp-mkp:.1f}%")
            if vb_hits:
                print(f"  🔥 Value Bet ▶ {'  '.join(vb_hits)}")

            # 差值
            if mkt_h:
                diffs = [f"主場{p_h-mkt_h:+.1f}%", f"平局{p_d-mkt_d:+.1f}%", f"客場{p_a-mkt_a:+.1f}%"]
                print(f"  ── 差值  ──  {'  '.join(diffs)}  （模型 - 市場，正=模型更看好）")

        # 收集輸出資料
        rows_out.append({
            "台灣日期": fix["tw_date"], "台灣時間": fix["tw_time"],
            "組別": fix["group"], "主隊": home, "客隊": away,
            # 賠率
            "主場賠率": odds_h, "平局賠率": odds_d, "客場賠率": odds_a,
            "大球(>2.5)": e.get("odds_over_25") if e else None,
            "小球(<2.5)": e.get("odds_under_25") if e else None,
            "雙隊進球": e.get("odds_btts_yes") if e else None,
            "莊家水份%": vig,
            # 市場機率
            "市場主場勝%": mkt_h, "市場平局%": mkt_d, "市場客場勝%": mkt_a,
            # 模型機率
            "模型主場勝%": pred_row["p_home"] if pred_row is not None else None,
            "模型平局%":   pred_row["p_draw"] if pred_row is not None else None,
            "模型客場勝%": pred_row["p_away"] if pred_row is not None else None,
            # 差值
            "差值主場%": round(float(pred_row["p_home"])-mkt_h,1) if pred_row is not None and mkt_h else None,
            "差值平局%": round(float(pred_row["p_draw"])-mkt_d,1) if pred_row is not None and mkt_d else None,
            "差值客場%": round(float(pred_row["p_away"])-mkt_a,1) if pred_row is not None and mkt_a else None,
            # 模型預測
            "模型預測": pred_row["pred_result"] if pred_row is not None else None,
            "預測主隊進球": int(pred_row["pred_home_goals"]) if pred_row is not None else None,
            "預測客隊進球": int(pred_row["pred_away_goals"]) if pred_row is not None else None,
            "#1比分機率%": pred_row["score1_prob"] if pred_row is not None else None,
            "#2比分": f"{int(pred_row['score2_home'])}-{int(pred_row['score2_away'])}" if pred_row is not None else None,
            "#3比分": f"{int(pred_row['score3_home'])}-{int(pred_row['score3_away'])}" if pred_row is not None else None,
            "Value Bet": "  ".join(vb_hits) if pred_row is not None and mkt_h else "",
            # 實際結果
            "實際比分": f"{int(pred_row['actual_home_score'])}-{int(pred_row['actual_away_score'])}" if pred_row is not None and str(pred_row.get("actual_result","")).strip() in ("H","D","A") else "",
            "實際結果": pred_row["actual_result"] if pred_row is not None else "",
            "預測正確": pred_row["correct"] if pred_row is not None else "",
        })

    # 輸出 Excel
    if export and rows_out:
        _export_odds_excel(rows_out, target)


def _export_odds_excel(rows, tw_date):
    """輸出賠率比較 Excel"""
    Path(ODDS_EXCEL).parent.mkdir(parents=True, exist_ok=True)

    # 若已有舊檔，讀入合併（避免覆蓋其他日期）
    new_df = pd.DataFrame(rows)
    if Path(ODDS_EXCEL).exists():
        try:
            old_df = pd.read_excel(ODDS_EXCEL)
            # 移除同日期的舊資料，換成新的
            old_df = old_df[old_df["台灣日期"] != tw_date]
            final_df = pd.concat([old_df, new_df], ignore_index=True)
        except Exception:
            final_df = new_df
    else:
        final_df = new_df

    final_df = final_df.sort_values(["台灣日期","台灣時間"]).reset_index(drop=True)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(ODDS_EXCEL, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="賠率比較")
            ws = writer.sheets["賠率比較"]

            thin   = Side(style="thin", color="BFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # 標題列
            for cell in ws[1]:
                cell.fill      = PatternFill("solid", fgColor="1F4E79")
                cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border
            ws.row_dimensions[1].height = 28

            col_names = [c.value for c in ws[1]]
            vb_idx = col_names.index("Value Bet")+1 if "Value Bet" in col_names else None
            pred_idx = col_names.index("模型預測")+1 if "模型預測" in col_names else None

            fill_vb   = PatternFill("solid", fgColor="FFF3CD")
            fill_home = PatternFill("solid", fgColor="E8F5E9")
            fill_away = PatternFill("solid", fgColor="FFF3E0")
            fill_draw = PatternFill("solid", fgColor="F3F3F3")
            fill_even = PatternFill("solid", fgColor="F5F8FF")

            for row_idx in range(2, ws.max_row+1):
                vb_val   = ws.cell(row=row_idx, column=vb_idx).value  if vb_idx   else ""
                pred_val = ws.cell(row=row_idx, column=pred_idx).value if pred_idx else ""
                if vb_val:
                    rf = fill_vb
                elif pred_val == "H":
                    rf = fill_home
                elif pred_val == "A":
                    rf = fill_away
                elif pred_val == "D":
                    rf = fill_draw
                else:
                    rf = fill_even if row_idx % 2 == 0 else None

                for cell in ws[row_idx]:
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font      = Font(size=9, name="Arial")
                    if rf: cell.fill = rf

            for i, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                ws.column_dimensions[get_column_letter(i)].width = 11
            for name, w in [("主隊",16),("客隊",16),("Value Bet",22)]:
                if name in col_names:
                    ws.column_dimensions[get_column_letter(col_names.index(name)+1)].width = w
            ws.freeze_panes = "A2"

        print(f"\n  💾 Excel 已更新：{ODDS_EXCEL}")

    except ImportError:
        csv_out = ODDS_EXCEL.replace(".xlsx",".csv")
        final_df.to_csv(csv_out, index=False, encoding="utf-8-sig")
        print(f"\n  💾 CSV 已儲存：{csv_out}")

# ════════════════════════════════════════════════════════════════
# 指令函式
# ════════════════════════════════════════════════════════════════

def _print_prediction(fix, pred, show_elo=True):
    home, away = fix["home"], fix["away"]
    winner = home if pred["pred_result"]=="H" else ("平局" if pred["pred_result"]=="D" else away)
    print(f"\n  台灣時間 {fix['tw_date']} {fix['tw_time']}  │  Group {fix['group']}")
    print(f"  {home:<22} vs  {away}")
    if show_elo:
        print(f"  ELO：{home} {pred['home_elo']}  vs  {away} {pred['away_elo']}")
    print(f"  主場勝 {pred['p_home']:>5.1f}%  │  平局 {pred['p_draw']:>5.1f}%  │  客場勝 {pred['p_away']:>5.1f}%")
    print(f"  預測結果 ▶ {winner}")
    print(f"  比分推薦 ▶ #{1} {pred['pred_home_goals']}-{pred['pred_away_goals']}({pred['score1_prob']}%)"
          f"  #{2} {pred['score2_home']}-{pred['score2_away']}({pred['score2_prob']}%)"
          f"  #{3} {pred['score3_home']}-{pred['score3_away']}({pred['score3_prob']}%)")
    if pred["motivation_note"]:
        print(pred["motivation_note"])


def cmd_predict(et_date_filter=None):
    """產出所有預測（用最新滾動狀態）"""
    print("📥 載入模型與特徵...")
    xgb_data, poi_data, ens_data = load_models()
    feat, h2h_dict = load_features()
    state = load_live_state()

    fixtures = WC2026_FIXTURES
    if et_date_filter:
        fixtures = [f for f in fixtures if f["et_date"] == et_date_filter]
        if not fixtures:
            print(f"❌ 找不到美東日期 {et_date_filter} 的比賽"); return

    pred_log = Path(PRED_LOG)
    log_df   = pd.read_csv(PRED_LOG) if pred_log.exists() else pd.DataFrame()
    # 只重算尚未有結果的比賽（已有實際比分的不覆蓋）
    if not log_df.empty:
        done = set(zip(
            log_df[log_df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]["home"],
            log_df[log_df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]["away"]
        ))
    else:
        done = set()

    label = f"美東 {et_date_filter}" if et_date_filter else "全部72場"
    print(f"\n{'='*65}")
    print(f"  2026 世界盃預測  [{label}]  （使用最新滾動狀態）")
    print(f"{'='*65}")

    new_rows, update_rows = [], []
    for fix in fixtures:
        home, away = fix["home"], fix["away"]
        if (home, away) in done:
            continue  # 已有真實結果的不重算

        pred = predict_match(home, away, fix["group"], xgb_data, poi_data, ens_data, feat, state, h2h_dict)
        _print_prediction(fix, pred)

        row_data = {
            "tw_date":fix["tw_date"],"tw_time":fix["tw_time"],"et_date":fix["et_date"],
            "group":fix["group"],"home":home,"away":away,
            "p_home":pred["p_home"],"p_draw":pred["p_draw"],"p_away":pred["p_away"],
            "pred_result":pred["pred_result"],
            "pred_home_goals":pred["pred_home_goals"],"pred_away_goals":pred["pred_away_goals"],
            "score1_prob":pred["score1_prob"],
            "score2_home":pred["score2_home"],"score2_away":pred["score2_away"],"score2_prob":pred["score2_prob"],
            "score3_home":pred["score3_home"],"score3_away":pred["score3_away"],"score3_prob":pred["score3_prob"],
            "home_elo":pred["home_elo"],"away_elo":pred["away_elo"],
            "actual_home_score":"","actual_away_score":"","actual_result":"","correct":"",
        }

        if not log_df.empty and ((log_df["home"]==home)&(log_df["away"]==away)).any():
            idx = log_df[(log_df["home"]==home)&(log_df["away"]==away)].index[0]
            for k,v in row_data.items():
                if k not in ("actual_home_score","actual_away_score","actual_result","correct"):
                    log_df.at[idx, k] = v
        else:
            new_rows.append(row_data)

    if new_rows:
        log_df = pd.concat([log_df, pd.DataFrame(new_rows)], ignore_index=True) if not log_df.empty else pd.DataFrame(new_rows)
    if not log_df.empty:
        log_df = log_df.sort_values(["tw_date","tw_time"]).reset_index(drop=True)
        log_df.to_csv(PRED_LOG, index=False, encoding="utf-8-sig")
        print(f"\n✅ 預測已儲存至 {PRED_LOG}")


def cmd_today(tw_date_filter=None):
    """查看今天台灣時間的比賽與最新預測"""
    target = tw_date_filter or (datetime.now(timezone.utc)+timedelta(hours=8)).strftime("%Y-%m-%d")
    fixtures_today = [f for f in WC2026_FIXTURES if f["tw_date"]==target]

    if not fixtures_today:
        print(f"\n台灣時間 {target} 沒有分組賽。")
        future = [f for f in WC2026_FIXTURES if f["tw_date"]>target]
        if future: print(f"下一個比賽日（台灣時間）：{future[0]['tw_date']}")
        return

    print(f"\n{'='*65}")
    print(f"  台灣時間 {target} 的比賽（共 {len(fixtures_today)} 場）")
    print(f"{'='*65}")

    pred_log = Path(PRED_LOG)
    log_df   = pd.read_csv(PRED_LOG) if pred_log.exists() else pd.DataFrame()

    # 一次性拉取賠率（避免每場都呼叫 API）
    if BSD_API_KEY and HAS_REQUESTS:
        bsd_events_all = _fetch_bsd_events()
        nm = BSD_NAME_MAP
        bsd_map_today = {
            (nm.get(e["home_team"],e["home_team"]), nm.get(e["away_team"],e["away_team"])): e
            for e in bsd_events_all
        }
    else:
        bsd_map_today = {}

    for fix in sorted(fixtures_today, key=lambda x: x["tw_time"]):
        home, away = fix["home"], fix["away"]
        print(f"\n  {fix['tw_time']}  Group {fix['group']}")
        print(f"  {home:<22} vs  {away}")

        if not log_df.empty:
            m = log_df[(log_df["home"]==home)&(log_df["away"]==away)]
            if not m.empty:
                r = m.iloc[0]
                winner = home if r["pred_result"]=="H" else ("平局" if r["pred_result"]=="D" else away)
                actual = ""
                if str(r.get("actual_result","")).strip() in ("H","D","A"):
                    actual = f"  ← 實際：{int(r['actual_home_score'])}-{int(r['actual_away_score'])} {'✅' if r['correct'] else '❌'}"
                elo_str = f"  ELO：{r.get('home_elo','?')} vs {r.get('away_elo','?')}"
                print(elo_str)
                print(f"  主場勝 {r['p_home']:>5.1f}%  │  平局 {r['p_draw']:>5.1f}%  │  客場勝 {r['p_away']:>5.1f}%")
                print(f"  預測 ▶ {winner}{actual}")
                print(f"  比分 ▶ #{1} {int(r['pred_home_goals'])}-{int(r['pred_away_goals'])}({r['score1_prob']}%)"
                      f"  #{2} {int(r['score2_home'])}-{int(r['score2_away'])}({r['score2_prob']}%)"
                      f"  #{3} {int(r['score3_home'])}-{int(r['score3_away'])}({r['score3_prob']}%)")
                # 賠率快速顯示（bsd_map 在 loop 外已建立，直接用）
                bsd_e = bsd_map_today.get((home, away)) if bsd_map_today else None
                if bsd_e and bsd_e.get("odds_home"):
                    mh,md,ma = _bsd_implied_prob(bsd_e["odds_home"],bsd_e["odds_draw"],bsd_e["odds_away"])
                    print(f"  賠率 ▶ H={bsd_e['odds_home']} D={bsd_e['odds_draw']} A={bsd_e['odds_away']}  市場機率：{mh}%/{md}%/{ma}%")
            else:
                print("  （尚未預測，執行 --predict）")
        else:
            print("  （尚未預測，執行 --predict）")


def _fetch_bsd_results(target_tw_date: str) -> dict:
    """
    從 BSD API 拉取指定台灣日期的 WC 完賽比分。
    回傳 dict：key=(home_mapped, away_mapped), value=(home_goals, away_goals)
    只回傳 status=finished 的比賽。
    """
    if not HAS_REQUESTS or not BSD_API_KEY:
        return {}
    headers = {"Authorization": f"Token {BSD_API_KEY}"}
    try:
        r = _requests.get(f"{BSD_BASE}/api/events/",
                          headers=headers,
                          params={"league": BSD_WC_LEAGUE_ID, "limit": 100},
                          timeout=10)
        if r.status_code != 200:
            return {}
        events = r.json().get("results", [])
    except Exception:
        return {}

    results = {}
    for e in events:
        if e.get("status") != "finished":
            continue
        if e.get("home_score") is None or e.get("away_score") is None:
            continue
        # 日期轉台灣時間
        event_dt = e.get("event_date","")
        try:
            dt_utc = datetime.fromisoformat(event_dt.replace("Z","+00:00"))
            dt_tw  = dt_utc.astimezone(timezone(timedelta(hours=8)))
            event_tw_date = dt_tw.strftime("%Y-%m-%d")
        except Exception:
            event_tw_date = event_dt[:10]

        if event_tw_date != target_tw_date:
            continue

        h = BSD_NAME_MAP.get(e["home_team"], e["home_team"])
        a = BSD_NAME_MAP.get(e["away_team"], e["away_team"])
        results[(h, a)] = (int(e["home_score"]), int(e["away_score"]))

    return results


def _process_result(df, idx, row, hg, ag, state, source_label):
    """比分寫入 df + 更新滾動狀態，回傳 (updated_state, elo_changes_str)"""
    actual_result = "H" if hg > ag else ("D" if hg == ag else "A")
    correct = (actual_result == row["pred_result"])

    df.at[idx, "actual_home_score"] = hg
    df.at[idx, "actual_away_score"] = ag
    df.at[idx, "actual_result"]     = actual_result
    df.at[idx, "correct"]           = correct

    home, away = row["home"], row["away"]
    elo_h_before = float(state[state["team"]==home]["elo"].iloc[0]) if not state[state["team"]==home].empty else 1700
    elo_a_before = float(state[state["team"]==away]["elo"].iloc[0]) if not state[state["team"]==away].empty else 1700

    state = update_live_state(state, home, away, hg, ag, row["group"])
    save_live_state(state)

    elo_h_after = float(state[state["team"]==home]["elo"].iloc[0])
    elo_a_after = float(state[state["team"]==away]["elo"].iloc[0])

    status = "✅ 預測正確！" if correct else "❌ 預測錯誤"
    print(f"  {status}  {source_label}：{hg}-{ag}")
    print(f"  ELO 更新：{home} {elo_h_before:.0f}→{elo_h_after:.0f}({elo_h_after-elo_h_before:+.0f})"
          f"  {away} {elo_a_before:.0f}→{elo_a_after:.0f}({elo_a_after-elo_a_before:+.0f})")
    return state, correct


def cmd_verify(tw_date_filter=None):
    """
    自動驗證流程：
    1. 先從 BSD API 拉已完賽的比分（status=finished）
    2. 有 API 比分 → 自動填入，不需手動
    3. 沒有 API 比分（比賽未結束）→ 詢問手動輸入
    4. 有任何更新 → 自動重算後續所有比賽預測
    """
    pred_log = Path(PRED_LOG)
    if not pred_log.exists():
        print("❌ 請先執行 --predict"); return

    target = tw_date_filter or (datetime.now(timezone.utc)+timedelta(hours=8)).strftime("%Y-%m-%d")
    df     = pd.read_csv(PRED_LOG, dtype={"actual_result":str})
    mask   = (df["tw_date"]==target) & (df["actual_result"].fillna("").str.strip()=="")
    unverified = df[mask]

    if unverified.empty:
        already = df[(df["tw_date"]==target) & (df["actual_result"].fillna("").str.strip().isin(["H","D","A"]))]
        if not already.empty:
            print(f"\n✅ 台灣時間 {target} 的比賽都已驗證完畢！")
        else:
            print(f"\n台灣時間 {target} 沒有待驗證的比賽（用 --date 指定日期）")
        cmd_stats(); return

    # ── 先從 BSD 拉已完賽比分 ──────────────────────────────────
    print(f"\n{'='*65}")
    print(f"  驗證台灣時間 {target} 的比賽結果")
    print(f"{'='*65}")

    if BSD_API_KEY and HAS_REQUESTS:
        print(f"\n  📡 從 BSD API 拉取已完賽比分...")
        bsd_results = _fetch_bsd_results(target)
        auto_count = len(bsd_results)
        if auto_count:
            print(f"  ✅ BSD 回傳 {auto_count} 場已完賽比分")
        else:
            print(f"  （BSD 目前無已完賽資料，將以手動輸入為主）")
    else:
        bsd_results = {}

    state = load_live_state()
    updated_any = False

    for idx, row in unverified.iterrows():
        home, away = row["home"], row["away"]
        print(f"\n  {row['tw_time']}  Group {row['group']}  {home} vs {away}")
        winner = home if row["pred_result"]=="H" else ("平局" if row["pred_result"]=="D" else away)
        print(f"  模型預測 ▶ {winner}  "
              f"（主場勝 {row['p_home']}% │ 平局 {row['p_draw']}% │ 客場勝 {row['p_away']}%）")
        print(f"  比分推薦 ▶ "
              f"#{1} {int(row['pred_home_goals'])}-{int(row['pred_away_goals'])}  "
              f"#{2} {int(row['score2_home'])}-{int(row['score2_away'])}  "
              f"#{3} {int(row['score3_home'])}-{int(row['score3_away'])}")

        # ── 情況一：BSD 有完賽比分 → 自動填入 ──────────────────
        if (home, away) in bsd_results:
            hg, ag = bsd_results[(home, away)]
            print(f"  🤖 BSD 自動取得比分：{hg}-{ag}")
            state, _ = _process_result(df, idx, row, hg, ag, state, "BSD自動")
            updated_any = True

        # ── 情況二：無 API 比分 → 手動輸入 ─────────────────────
        else:
            print(f"  （BSD 尚無完賽資料）")
            score_input = input("  手動輸入比分（格式 2-1，Enter 跳過）：").strip()
            if not score_input:
                print("  ↩ 跳過（比賽尚未結束）")
                continue
            try:
                hg, ag = map(int, score_input.split("-"))
                state, _ = _process_result(df, idx, row, hg, ag, state, "手動輸入")
                updated_any = True
            except ValueError:
                print("  ⚠ 格式錯誤，已跳過")

    df.to_csv(PRED_LOG, index=False, encoding="utf-8-sig")

    # ── 有更新 → 重算後續預測 ────────────────────────────────
    if updated_any:
        print(f"\n🔄 重算後續比賽預測（ELO/form 已更新）...")
        xgb_data, poi_data, ens_data = load_models()
        feat, h2h_dict = load_features()
        state = load_live_state()

        done_keys = set(zip(
            df[df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]["home"],
            df[df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]["away"]
        ))
        recalc_count = 0
        for fix in WC2026_FIXTURES:
            home, away = fix["home"], fix["away"]
            if (home, away) in done_keys:
                continue
            pred = predict_match(home, away, fix["group"],
                                 xgb_data, poi_data, ens_data, feat, state, h2h_dict)
            mask2 = (df["home"]==home) & (df["away"]==away)
            if mask2.any():
                idx2 = df[mask2].index[0]
                for k in ["p_home","p_draw","p_away","pred_result",
                           "pred_home_goals","pred_away_goals","score1_prob",
                           "score2_home","score2_away","score2_prob",
                           "score3_home","score3_away","score3_prob",
                           "home_elo","away_elo"]:
                    df.at[idx2, k] = pred[k]
            recalc_count += 1

        df = df.sort_values(["tw_date","tw_time"]).reset_index(drop=True)
        df.to_csv(PRED_LOG, index=False, encoding="utf-8-sig")
        print(f"  ✅ {recalc_count} 場後續預測已自動更新")
    else:
        print(f"\n  沒有新的比賽結果，預測維持不變")

    cmd_stats()


def cmd_stats():
    pred_log = Path(PRED_LOG)
    if not pred_log.exists():
        print("❌ 找不到預測記錄"); return

    df = pd.read_csv(PRED_LOG, dtype={"actual_result":str})
    verified = df[df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]

    print(f"\n{'='*55}")
    print(f"  模型準確率統計")
    print(f"{'='*55}")

    if verified.empty:
        print("  尚無已驗證的比賽")
        print(f"  待驗證：{len(df)} 場")
        print(f"{'='*55}"); return

    total   = len(verified)
    correct = int(verified["correct"].sum())
    print(f"  已驗證：{total} 場  │  正確：{correct} 場  │  準確率：{correct/total*100:.1f}%")

    print(f"\n  ── 各日明細 ──")
    verified_sorted = verified.sort_values(["tw_date","tw_time"]).reset_index(drop=True)
    for tw_date, grp in verified_sorted.groupby("tw_date", sort=False):
        dc = int(grp["correct"].sum())
        print(f"\n  {tw_date}  {dc}/{len(grp)} 場正確")
        for _, r in grp.sort_values("tw_time").iterrows():
            st = "✅" if r["correct"] else "❌"
            tw_time = r.get("tw_time","")
            print(f"    {st} {tw_time}  {r['home']:<20} {int(r['actual_home_score'])}-{int(r['actual_away_score'])} {r['away']:<20}  (預測 {r['pred_result']})")

    print(f"\n  ── 各結果類型 ──")
    for res, label in [("H","主場勝"),("D","平局"),("A","客場勝")]:
        sub = verified[verified["actual_result"]==res]
        if len(sub):
            sc = int(sub["correct"].sum())
            print(f"  {label}：{sc}/{len(sub)} ({sc/len(sub)*100:.0f}%)")

    unv = df[~df["actual_result"].fillna("").str.strip().isin(["H","D","A"])]
    print(f"\n  待驗證：{len(unv)} 場")
    print(f"{'='*55}")


def cmd_standings(group_filter=None):
    """查看目前積分榜"""
    state = load_live_state()
    groups = [group_filter] if group_filter else sorted(state["group"].unique())

    print(f"\n{'='*55}")
    print(f"  目前積分榜")
    print(f"{'='*55}")
    for g in groups:
        standings = get_group_standings(state, g)
        print(f"\n  Group {g}")
        print(f"  {'隊伍':<22} {'積分':>4} {'場':>3} {'勝':>3} {'平':>3} {'負':>3} {'進':>3} {'失':>3} {'ELO':>6}")
        print(f"  {'-'*52}")
        for i, row in standings.iterrows():
            rank_sym = "▲" if i < 2 else " "
            print(f"  {rank_sym}{row['team']:<21} {int(row['pts']):>4} {int(row['played']):>3} "
                  f"{int(row['w']):>3} {int(row['d']):>3} {int(row['l']):>3} "
                  f"{int(row['gf']):>3} {int(row['ga']):>3} {row['elo']:>6.0f}")
    print(f"\n  ▲ = 目前積分前2名（暫時晉級位置）")
    print(f"{'='*55}")


# ════════════════════════════════════════════════════════════════
# 主程式
# ════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="2026 世界盃預測驗證工具（滾動更新版）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
每日標準流程：
  1. 早上查看今天比賽  →  python src/verify_predictions.py --today
  2. 賽後輸入結果      →  python src/verify_predictions.py --verify
     ↑ 自動更新 ELO/form/積分，並重算所有後續預測
  3. 查看積分榜        →  python src/verify_predictions.py --standings
  4. 查看累計成績      →  python src/verify_predictions.py --stats

搭配 --date 使用（台灣日期）：
  --today     --date 2026-06-14
  --verify    --date 2026-06-14
  --standings --group A
        """
    )
    parser.add_argument("--predict",    action="store_true", help="產出所有預測（使用最新滾動狀態）")
    parser.add_argument("--today",      action="store_true", help="查看今天台灣時間的比賽")
    parser.add_argument("--verify",     action="store_true", help="輸入比分並更新後續預測")
    parser.add_argument("--odds",       action="store_true", help="拉取 BSD 賠率並與模型比對（輸出 Excel）")
    parser.add_argument("--stats",      action="store_true", help="查看準確率統計")
    parser.add_argument("--standings",  action="store_true", help="查看積分榜")
    parser.add_argument("--date",       type=str, help="台灣日期 YYYY-MM-DD")
    parser.add_argument("--group",      type=str, help="指定組別（A-L），搭配 --standings 使用")
    args = parser.parse_args()

    if   args.predict:   cmd_predict(et_date_filter=args.date)
    elif args.today:     cmd_today(tw_date_filter=args.date)
    elif args.verify:    cmd_verify(tw_date_filter=args.date)
    elif args.odds:      cmd_odds(tw_date_filter=args.date)
    elif args.stats:     cmd_stats()
    elif args.standings: cmd_standings(group_filter=args.group)
    else:
        print(__doc__)
        print("\n提示：執行 python src/verify_predictions.py --help 查看所有指令")