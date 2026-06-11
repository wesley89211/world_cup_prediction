"""
fetch_bsd_odds.py
=================
從 BSD API 拉取世界盃賠率，與模型預測機率比對，
偵測 Value Bet 並輸出到 Excel。

執行：
  python src/fetch_bsd_odds.py

輸出：
  data/reports/wc2026_odds_comparison.xlsx

功能：
  1. 拉取 World Cup 2026 所有比賽賠率（league_id=27）
  2. 賠率轉換成隱含機率（去除莊家水份）
  3. 與模型預測機率比對
  4. 標記 Value Bet（模型機率 > 市場機率 + 閾值）
  5. 輸出格式化 Excel
"""

import requests
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timezone, timedelta

# ── 設定 ──────────────────────────────────────────────────────
BSD_API_KEY  = "6c8388e0a102b8f1c3e160ddda55310b802dd003"
BSD_BASE     = "https://sports.bzzoiro.com"
WC_LEAGUE_ID = 27           # World Cup 2026

PRED_LOG     = r"data/processed/wc2026_predictions.csv"
OUTPUT_EXCEL = r"data/reports/wc2026_odds_comparison.xlsx"

VALUE_BET_THRESHOLD = 0.05  # 模型機率超過市場機率 5% 以上才標記
# ─────────────────────────────────────────────────────────────

HEADERS = {"Authorization": f"Token {BSD_API_KEY}"}


# ════════════════════════════════════════════════════════════════
# BSD API 函式
# ════════════════════════════════════════════════════════════════

def fetch_wc_odds() -> pd.DataFrame:
    """拉取 World Cup 2026 所有比賽賠率"""
    print("📡 拉取 BSD 世界盃賠率...")

    all_events = []
    url = f"{BSD_BASE}/api/events/"
    params = {"league": WC_LEAGUE_ID, "limit": 100, "offset": 0}

    while url:
        r = requests.get(url, headers=HEADERS, params=params)
        if r.status_code != 200:
            print(f"  ❌ API 錯誤：{r.status_code} {r.text[:200]}")
            break
        data = r.json()
        all_events.extend(data.get("results", []))
        url    = data.get("next")
        params = {}  # next URL 已含 params

    print(f"  共取得 {len(all_events)} 場比賽")

    rows = []
    for e in all_events:
        # 日期轉台灣時間
        event_dt = e.get("event_date", "")
        if event_dt:
            try:
                dt_utc = datetime.fromisoformat(event_dt.replace("Z", "+00:00"))
                dt_tw  = dt_utc.astimezone(timezone(timedelta(hours=8)))
                tw_date = dt_tw.strftime("%Y-%m-%d")
                tw_time = dt_tw.strftime("%H:%M")
            except Exception:
                tw_date = event_dt[:10]
                tw_time = ""
        else:
            tw_date = tw_time = ""

        rows.append({
            "bsd_id":       e["id"],
            "tw_date":      tw_date,
            "tw_time":      tw_time,
            "home":         e["home_team"],
            "away":         e["away_team"],
            "status":       e.get("status", ""),
            "home_score":   e.get("home_score"),
            "away_score":   e.get("away_score"),
            # 賠率（decimal）
            "odds_home":    e.get("odds_home"),
            "odds_draw":    e.get("odds_draw"),
            "odds_away":    e.get("odds_away"),
            "odds_over25":  e.get("odds_over_25"),
            "odds_under25": e.get("odds_under_25"),
            "odds_btts_y":  e.get("odds_btts_yes"),
            "odds_btts_n":  e.get("odds_btts_no"),
        })

    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════════
# 賠率計算函式
# ════════════════════════════════════════════════════════════════

def implied_prob(odds_h, odds_d, odds_a):
    """
    將 decimal 賠率轉換成無水份隱含機率。
    vig（莊家水份）= 1/H + 1/D + 1/A - 1
    去水份後：p = (1/odds) / (1/H + 1/D + 1/A)
    """
    if not all([odds_h, odds_d, odds_a]):
        return None, None, None
    try:
        raw_h = 1 / float(odds_h)
        raw_d = 1 / float(odds_d)
        raw_a = 1 / float(odds_a)
        total = raw_h + raw_d + raw_a  # > 1，差值就是 vig
        return round(raw_h/total*100, 1), round(raw_d/total*100, 1), round(raw_a/total*100, 1)
    except (ZeroDivisionError, TypeError):
        return None, None, None


def calc_vig(odds_h, odds_d, odds_a):
    """計算莊家水份（%）"""
    try:
        total = 1/float(odds_h) + 1/float(odds_d) + 1/float(odds_a)
        return round((total - 1) * 100, 2)
    except (TypeError, ZeroDivisionError):
        return None


def best_odds_result(odds_h, odds_d, odds_a):
    """賠率最高（市場最不看好）的結果"""
    mapping = {"主場勝": odds_h, "平局": odds_d, "客場勝": odds_a}
    valid = {k: v for k, v in mapping.items() if v}
    return max(valid, key=valid.get) if valid else ""


# ════════════════════════════════════════════════════════════════
# 主程式
# ════════════════════════════════════════════════════════════════

def main():
    # ── 拉取賠率 ──
    odds_df = fetch_wc_odds()

    # ── 計算隱含機率 ──
    odds_df[["mkt_p_home","mkt_p_draw","mkt_p_away"]] = odds_df.apply(
        lambda r: pd.Series(implied_prob(r["odds_home"], r["odds_draw"], r["odds_away"])),
        axis=1
    )
    odds_df["vig_pct"] = odds_df.apply(
        lambda r: calc_vig(r["odds_home"], r["odds_draw"], r["odds_away"]), axis=1
    )

    # ── 讀取模型預測 ──
    pred_path = Path(PRED_LOG)
    has_model = pred_path.exists()

    if has_model:
        print("\n📊 讀取模型預測...")
        pred_df = pd.read_csv(PRED_LOG)

        # 隊名對齊（BSD 用 Czechia，我們用 Czech Republic 等）
        name_map = {
            "Czechia":        "Czech Republic",
            "Türkiye":        "Turkey",
            "Côte d'Ivoire":  "Ivory Coast",
            "Cabo Verde":     "Cape Verde",
            "USA":            "United States",
        }
        odds_df["home_mapped"] = odds_df["home"].replace(name_map)
        odds_df["away_mapped"] = odds_df["away"].replace(name_map)

        # 合併
        merged = odds_df.merge(
            pred_df[["home","away","p_home","p_draw","p_away","pred_result",
                     "pred_home_goals","pred_away_goals","score1_prob",
                     "score2_home","score2_away","score2_prob",
                     "score3_home","score3_away","score3_prob",
                     "actual_home_score","actual_away_score","actual_result","correct"]],
            left_on=["home_mapped","away_mapped"],
            right_on=["home","away"],
            how="left",
            suffixes=("","_pred")
        )
    else:
        print("\n⚠️  找不到預測記錄，只輸出賠率資料")
        merged = odds_df.copy()
        merged["p_home"] = merged["p_draw"] = merged["p_away"] = None
        merged["pred_result"] = None

    # ── Value Bet 偵測 ──
    def detect_value(row):
        results = []
        for result, model_p, mkt_p in [
            ("主場勝", row.get("p_home"), row.get("mkt_p_home")),
            ("平局",   row.get("p_draw"), row.get("mkt_p_draw")),
            ("客場勝", row.get("p_away"), row.get("mkt_p_away")),
        ]:
            if model_p and mkt_p:
                diff = float(model_p) - float(mkt_p)
                if diff >= VALUE_BET_THRESHOLD * 100:
                    results.append(f"{result}+{diff:.1f}%")
        return " | ".join(results) if results else ""

    merged["value_bet"] = merged.apply(detect_value, axis=1)

    # ── 差值欄位 ──
    for side, mp, mkp in [("home","p_home","mkt_p_home"),
                           ("draw","p_draw","mkt_p_draw"),
                           ("away","p_away","mkt_p_away")]:
        merged[f"diff_{side}"] = merged.apply(
            lambda r: round(float(r[mp]) - float(r[mkp]), 1)
            if pd.notna(r.get(mp)) and pd.notna(r.get(mkp)) else None,
            axis=1
        )

    # ── 整理輸出欄位 ──
    has_score = "actual_home_score" in merged.columns

    output_cols = [
        "tw_date", "tw_time", "home", "away", "status",
        # 賠率
        "odds_home", "odds_draw", "odds_away",
        "odds_over25", "odds_under25", "odds_btts_y",
        # 莊家水份
        "vig_pct",
        # 市場隱含機率
        "mkt_p_home", "mkt_p_draw", "mkt_p_away",
        # 模型機率
        "p_home", "p_draw", "p_away",
        # 差值（模型 - 市場）
        "diff_home", "diff_draw", "diff_away",
        # 模型預測
        "pred_result", "pred_home_goals", "pred_away_goals",
        "score1_prob",
        "score2_home", "score2_away", "score2_prob",
        "score3_home", "score3_away", "score3_prob",
        # Value Bet
        "value_bet",
    ]
    if has_score:
        output_cols += ["actual_home_score","actual_away_score","actual_result","correct"]

    # 只保留存在的欄位
    output_cols = [c for c in output_cols if c in merged.columns]
    out = merged[output_cols].sort_values(["tw_date","tw_time"]).reset_index(drop=True)

    # ── 終端機預覽 ──
    print(f"\n{'='*65}")
    print(f"  2026 世界盃賠率 vs 模型預測比較")
    print(f"{'='*65}")

    has_odds = out[out["odds_home"].notna()]
    print(f"\n有賠率的比賽：{len(has_odds)} 場 / 共 {len(out)} 場")

    for _, r in has_odds.head(20).iterrows():
        score_str = ""
        if has_score and pd.notna(r.get("actual_home_score")):
            score_str = f" ← 實際 {int(r['actual_home_score'])}-{int(r['actual_away_score'])}"
        print(f"\n  {r['tw_date']} {r['tw_time']}  {r['home']} vs {r['away']}{score_str}")
        print(f"  賠率   ：H={r['odds_home']}  D={r['odds_draw']}  A={r['odds_away']}  (水份={r['vig_pct']}%)")
        print(f"  市場機率：主場勝 {r['mkt_p_home']}%  平局 {r['mkt_p_draw']}%  客場勝 {r['mkt_p_away']}%")
        if pd.notna(r.get("p_home")):
            print(f"  模型機率：主場勝 {r['p_home']}%  平局 {r['p_draw']}%  客場勝 {r['p_away']}%")
            print(f"  差值    ：主場 {r['diff_home']:+.1f}%  平局 {r['diff_draw']:+.1f}%  客場 {r['diff_away']:+.1f}%")
        if r.get("value_bet"):
            print(f"  🔥 Value Bet：{r['value_bet']}")
        if pd.notna(r.get("pred_home_goals")):
            print(f"  比分推薦：#{1} {int(r['pred_home_goals'])}-{int(r['pred_away_goals'])}({r['score1_prob']}%)"
                  f"  #{2} {int(r['score2_home'])}-{int(r['score2_away'])}({r['score2_prob']}%)"
                  f"  #{3} {int(r['score3_home'])}-{int(r['score3_away'])}({r['score3_prob']}%)")

    # ── 輸出 Excel ──
    print(f"\n💾 輸出 Excel...")
    Path(OUTPUT_EXCEL).parent.mkdir(parents=True, exist_ok=True)

    # 欄位中文名對照
    col_rename = {
        "tw_date":          "台灣日期",
        "tw_time":          "台灣時間",
        "home":             "主隊",
        "away":             "客隊",
        "status":           "狀態",
        "odds_home":        "主場賠率",
        "odds_draw":        "平局賠率",
        "odds_away":        "客場賠率",
        "odds_over25":      "大球(>2.5)賠率",
        "odds_under25":     "小球(<2.5)賠率",
        "odds_btts_y":      "雙隊進球賠率",
        "vig_pct":          "莊家水份%",
        "mkt_p_home":       "市場主場勝%",
        "mkt_p_draw":       "市場平局%",
        "mkt_p_away":       "市場客場勝%",
        "p_home":           "模型主場勝%",
        "p_draw":           "模型平局%",
        "p_away":           "模型客場勝%",
        "diff_home":        "差值主場%",
        "diff_draw":        "差值平局%",
        "diff_away":        "差值客場%",
        "pred_result":      "模型預測",
        "pred_home_goals":  "預測主隊進球",
        "pred_away_goals":  "預測客隊進球",
        "score1_prob":      "#1比分機率%",
        "score2_home":      "#2比分主",
        "score2_away":      "#2比分客",
        "score2_prob":      "#2機率%",
        "score3_home":      "#3比分主",
        "score3_away":      "#3比分客",
        "score3_prob":      "#3機率%",
        "value_bet":        "Value Bet",
        "actual_home_score":"實際主隊進球",
        "actual_away_score":"實際客隊進球",
        "actual_result":    "實際結果",
        "correct":          "預測正確",
    }
    out_renamed = out.rename(columns=col_rename)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
            out_renamed.to_excel(writer, index=False, sheet_name="賠率比較")
            ws = writer.sheets["賠率比較"]

            thin   = Side(style="thin", color="BFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=False)

            # 標題列
            for cell in ws[1]:
                cell.fill      = PatternFill("solid", fgColor="1F4E79")
                cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Arial")
                cell.alignment = center
                cell.border    = border
            ws.row_dimensions[1].height = 28

            # 找各欄 index
            col_names = [c.value for c in ws[1]]
            def ci(name):
                return col_names.index(name) + 1 if name in col_names else None

            # 資料列
            fill_vb     = PatternFill("solid", fgColor="FFF3CD")  # Value Bet → 黃
            fill_even   = PatternFill("solid", fgColor="F5F8FF")
            fill_home_w = PatternFill("solid", fgColor="E8F5E9")  # 模型主場勝
            fill_away_w = PatternFill("solid", fgColor="FFF3E0")  # 模型客場勝
            fill_draw_w = PatternFill("solid", fgColor="F3F3F3")  # 模型平局

            pred_col = ci("模型預測")
            vb_col   = ci("Value Bet")

            for row_idx in range(2, ws.max_row + 1):
                pred_val = ws.cell(row=row_idx, column=pred_col).value if pred_col else ""
                vb_val   = ws.cell(row=row_idx, column=vb_col).value  if vb_col  else ""

                if vb_val:
                    row_fill = fill_vb
                elif pred_val == "H":
                    row_fill = fill_home_w
                elif pred_val == "A":
                    row_fill = fill_away_w
                elif pred_val == "D":
                    row_fill = fill_draw_w
                else:
                    row_fill = fill_even if row_idx % 2 == 0 else None

                for cell in ws[row_idx]:
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font      = Font(size=9, name="Arial")
                    if row_fill:
                        cell.fill = row_fill

            # 欄寬
            for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                ws.column_dimensions[get_column_letter(i)].width = 12

            # 特定欄加寬
            for name, width in [("主隊",18),("客隊",18),("Value Bet",22),
                                 ("台灣日期",12),("台灣時間",8)]:
                idx = ci(name)
                if idx:
                    ws.column_dimensions[get_column_letter(idx)].width = width

            ws.freeze_panes = "A2"

        print(f"✅ 已儲存至 {OUTPUT_EXCEL}")

    except ImportError:
        csv_out = OUTPUT_EXCEL.replace(".xlsx", ".csv")
        out_renamed.to_csv(csv_out, index=False, encoding="utf-8-sig")
        print(f"⚠️  未安裝 openpyxl，已存為 CSV：{csv_out}")
        print("    安裝方式：pip install openpyxl")

    # ── 統計摘要 ──
    vb_count = (out["value_bet"].notna() & (out["value_bet"] != "")).sum()
    print(f"\n📊 摘要：")
    print(f"  總比賽場次  ：{len(out)}")
    print(f"  有賠率場次  ：{out['odds_home'].notna().sum()}")
    print(f"  Value Bet   ：{vb_count} 場")
    if vb_count:
        print(f"\n  🔥 Value Bet 清單：")
        for _, r in out[out["value_bet"].notna() & (out["value_bet"] != "")].iterrows():
            print(f"    {r['tw_date']} {r['home']} vs {r['away']} → {r['value_bet']}")


if __name__ == "__main__":
    main()
