"""
predict_matchup.py
==================
手動輸入任意兩隊對戰，預測結果並輸出到 Excel。

使用方式：
  python src/predict_matchup.py

功能：
  - 互動式選隊（輸入編號或隊名，支援模糊搜尋）
  - 單場預測：勝/平/負機率 + 比分推薦前三
  - 可連續預測多場
  - 所有預測結果累積輸出到 Excel（含格式）
  - Excel 路徑：data/reports/matchup_predictions.xlsx
"""

import pandas as pd
import numpy as np
import pickle
from pathlib import Path
from datetime import datetime
from scipy.stats import poisson as scipy_poisson
import difflib

# ── 路徑設定 ─────────────────────────────────────────────────────
XGB_PATH      = r"data/models/xgb_v2.pkl"
POISSON_PATH  = r"data/models/advanced_poisson_v2.pkl"
ENSEMBLE_PATH = r"data/models/ensemble_v2.pkl"
FEATURES_PATH = r"data/processed/wc26_final_features.csv"
ELO_PATH      = r"data/processed/wc26_elo_v2.csv"
H2H_PATH      = r"data/processed/wc26_h2h.csv"
LIVE_STATE    = r"data/processed/wc2026_live_state.csv"
OUTPUT_EXCEL  = r"data/reports/matchup_predictions.xlsx"
# ─────────────────────────────────────────────────────────────────

WC_GOAL_SCALE = 1.35  # 世界盃進球校正係數


# ════════════════════════════════════════════════════════════════
# 載入模型與資料
# ════════════════════════════════════════════════════════════════

def load_all():
    print("📥 載入模型與資料...")

    with open(XGB_PATH,      "rb") as f: xgb_data  = pickle.load(f)
    with open(POISSON_PATH,  "rb") as f: poi_data  = pickle.load(f)
    with open(ENSEMBLE_PATH, "rb") as f: ens_data  = pickle.load(f)

    feat = pd.read_csv(FEATURES_PATH).set_index("team")

    # ELO：優先用滾動狀態（比賽中更新過的），否則用初始值
    if Path(LIVE_STATE).exists():
        live = pd.read_csv(LIVE_STATE).set_index("team")
        elo  = live["elo"].to_dict()
    else:
        elo_df = pd.read_csv(ELO_PATH)
        elo    = dict(zip(elo_df["team"], elo_df["elo_weighted"]))

    # H2H
    h2h_raw  = pd.read_csv(H2H_PATH)
    h2h_dict = {}
    for _, r in h2h_raw.iterrows():
        a, b, t = r["team_a"], r["team_b"], r["total_matches"]
        if t == 0: continue
        h2h_dict[(a,b)] = {"h2h_total":t,"h2h_home_winrate":r["team_a_wins"]/t,"h2h_draw_rate":r["draws"]/t}
        h2h_dict[(b,a)] = {"h2h_total":t,"h2h_home_winrate":r["team_b_wins"]/t,"h2h_draw_rate":r["draws"]/t}

    teams = sorted(feat.index.tolist())
    print(f"✅ 就緒（共 {len(teams)} 支隊伍，ELO 來源：{'滾動更新版' if Path(LIVE_STATE).exists() else '初始版'}）\n")
    return xgb_data, poi_data, ens_data, feat, elo, h2h_dict, teams


# ════════════════════════════════════════════════════════════════
# 特徵建立 & 預測
# ════════════════════════════════════════════════════════════════

def build_features(home, away, feat, elo, h2h_dict):
    feat_means  = feat.mean()
    ELO_DEFAULT = 1700

    def get_f(team, col):
        if team in feat.index:
            v = feat.at[team, col]
            return v if pd.notna(v) else feat_means[col]
        return feat_means[col]

    he  = elo.get(home, ELO_DEFAULT)
    ae  = elo.get(away, ELO_DEFAULT)
    h2h = h2h_dict.get((home, away), {"h2h_total":0,"h2h_home_winrate":0.333,"h2h_draw_rate":0.25})

    row = {
        "home_elo":he, "away_elo":ae, "elo_diff":he-ae,
        "is_neutral":1, "is_wc_final":1, "is_friendly":0,
        "h2h_home_winrate":h2h["h2h_home_winrate"],
        "h2h_draw_rate":   h2h["h2h_draw_rate"],
        "h2h_total":       h2h["h2h_total"],
    }
    for col in feat.columns:
        row[f"home_{col}"] = get_f(home, col)
        row[f"away_{col}"] = get_f(away, col)

    diff_targets = ["total_value_eur","game_top_11_ovr","game_fw_speed","game_fw_finishing",
                    "game_mf_passing","game_df_defense","game_df_physic","win_rate_r5","win_rate_wc",
                    "goal_diff_avg_r5","goals_for_avg_r5","goals_against_avg_r5",
                    "form_score_official","top_club_ratio","avg_age"]
    for col in diff_targets:
        if f"home_{col}" in row:
            row[f"diff_{col}"] = row[f"home_{col}"] - row[f"away_{col}"]
    row["diff_value_M"] = row.get("diff_total_value_eur", 0) / 1e6
    return row, he, ae


def poisson_probs(lh, la, max_goals=8):
    p_h = p_d = p_a = 0.0
    for gh in range(max_goals+1):
        for ga in range(max_goals+1):
            p = scipy_poisson.pmf(gh,lh) * scipy_poisson.pmf(ga,la)
            if   gh > ga:  p_h += p
            elif gh == ga: p_d += p
            else:          p_a += p
    t = p_h + p_d + p_a
    return p_h/t, p_d/t, p_a/t


def predict(home, away, xgb_data, poi_data, ens_data, feat, elo, h2h_dict):
    row, he, ae = build_features(home, away, feat, elo, h2h_dict)

    # XGBoost
    X_xgb    = np.array([[row.get(f,0) for f in xgb_data["features"]]])
    xgb_prob = xgb_data["model"].predict_proba(X_xgb)[0]  # [away, draw, home]

    # Poisson
    X_poi   = np.array([[row.get(f,0) for f in poi_data["poi_cols"]]])
    X_poi_s = poi_data["scaler"].transform(X_poi)
    lh_raw  = max(poi_data["pr_home"].predict(X_poi_s)[0], 0.05)
    la_raw  = max(poi_data["pr_away"].predict(X_poi_s)[0], 0.05)
    ph, pd_, pa = poisson_probs(lh_raw, la_raw)
    poi_prob = np.array([pa, pd_, ph])

    # Ensemble
    w        = ens_data["xgb_weight"]
    ens_prob = w * xgb_prob + (1-w) * poi_prob  # [away, draw, home]

    # 比分 Top 5（WC 校正）
    lh = lh_raw * WC_GOAL_SCALE
    la = la_raw * WC_GOAL_SCALE
    sp = [(gh, ga, round(scipy_poisson.pmf(gh,lh)*scipy_poisson.pmf(ga,la)*100, 1))
          for gh in range(11) for ga in range(11)]
    sp.sort(key=lambda x: -x[2])
    top5 = sp[:5]

    pred_result = ("主場勝" if ens_prob[2] > ens_prob[0] and ens_prob[2] > ens_prob[1]
                   else ("平局" if ens_prob[1] >= ens_prob[0] and ens_prob[1] >= ens_prob[2]
                         else "客場勝"))
    winner = home if pred_result=="主場勝" else (away if pred_result=="客場勝" else "平局")

    return {
        "home": home, "away": away,
        "home_elo": round(he, 0), "away_elo": round(ae, 0),
        "p_home": round(ens_prob[2]*100, 1),
        "p_draw": round(ens_prob[1]*100, 1),
        "p_away": round(ens_prob[0]*100, 1),
        "pred_result": pred_result,
        "winner": winner,
        "score1": f"{top5[0][0]}-{top5[0][1]}", "score1_prob": top5[0][2],
        "score2": f"{top5[1][0]}-{top5[1][1]}", "score2_prob": top5[1][2],
        "score3": f"{top5[2][0]}-{top5[2][1]}", "score3_prob": top5[2][2],
        "score4": f"{top5[3][0]}-{top5[3][1]}", "score4_prob": top5[3][2],
        "score5": f"{top5[4][0]}-{top5[4][1]}", "score5_prob": top5[4][2],
        "lambda_home": round(lh, 2), "lambda_away": round(la, 2),
        "h2h_matches": int(h2h_dict.get((home,away),{}).get("h2h_total",0)),
        "h2h_home_wr": round(h2h_dict.get((home,away),{}).get("h2h_home_winrate",0.333)*100, 1),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ════════════════════════════════════════════════════════════════
# 互動式選隊
# ════════════════════════════════════════════════════════════════

def print_team_list(teams):
    print("\n" + "─"*55)
    print("  48 支參賽隊伍：")
    print("─"*55)
    for i, t in enumerate(teams):
        end = "\n" if (i+1) % 3 == 0 else ""
        print(f"  {i+1:2d}. {t:<22}", end=end)
    if len(teams) % 3 != 0:
        print()
    print("─"*55)


def fuzzy_find_team(query, teams):
    """模糊比對隊名，回傳最接近的隊伍"""
    query = query.strip().lower()

    # 直接完整比對
    for t in teams:
        if t.lower() == query:
            return t

    # 前綴比對
    matches = [t for t in teams if t.lower().startswith(query)]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        return matches  # 回傳多個候選

    # 模糊比對
    close = difflib.get_close_matches(query, [t.lower() for t in teams], n=3, cutoff=0.5)
    if close:
        found = [t for t in teams if t.lower() in close]
        if len(found) == 1:
            return found[0]
        return found

    return None


def select_team(teams, prompt, exclude=None):
    """互動選隊，支援編號或隊名輸入"""
    while True:
        raw = input(f"\n  {prompt}（輸入編號或隊名，? 列出所有）：").strip()

        if raw == "?":
            print_team_list(teams)
            continue

        if raw == "":
            print("  ⚠ 請輸入編號或隊名")
            continue

        # 編號輸入
        if raw.isdigit():
            idx = int(raw) - 1
            if 0 <= idx < len(teams):
                team = teams[idx]
                if exclude and team == exclude:
                    print(f"  ⚠ 不能選同一支隊伍")
                    continue
                print(f"  ✅ 選擇：{team}")
                return team
            else:
                print(f"  ⚠ 請輸入 1~{len(teams)} 之間的編號")
                continue

        # 文字輸入（模糊比對）
        result = fuzzy_find_team(raw, teams)

        if result is None:
            print(f"  ⚠ 找不到「{raw}」，請重試（? 列出所有隊伍）")
            continue

        if isinstance(result, list):
            print(f"  找到多個符合的隊伍：")
            for i, t in enumerate(result):
                print(f"    {i+1}. {t}")
            sub = input("  請選擇編號：").strip()
            if sub.isdigit() and 1 <= int(sub) <= len(result):
                team = result[int(sub)-1]
                if exclude and team == exclude:
                    print(f"  ⚠ 不能選同一支隊伍")
                    continue
                print(f"  ✅ 選擇：{team}")
                return team
            continue

        # 單一比對結果
        if exclude and result == exclude:
            print(f"  ⚠ 不能選同一支隊伍")
            continue
        print(f"  ✅ 選擇：{result}")
        return result


# ════════════════════════════════════════════════════════════════
# 顯示預測結果
# ════════════════════════════════════════════════════════════════

def print_result(r):
    SEP = "─" * 55
    print(f"\n  {SEP}")
    print(f"  {r['home']:<22}  vs  {r['away']}")
    print(f"  ELO：{r['home']} {r['home_elo']:.0f}  vs  {r['away']} {r['away_elo']:.0f}")
    print(f"  {SEP}")
    print(f"  主場勝  {r['p_home']:>5.1f}%  │  平局  {r['p_draw']:>5.1f}%  │  客場勝  {r['p_away']:>5.1f}%")
    print(f"  預測結果  ▶  {r['winner']}")
    print(f"  {SEP}")
    print(f"  比分推薦：")
    for i, (s, sp) in enumerate([
        (r["score1"],r["score1_prob"]),(r["score2"],r["score2_prob"]),
        (r["score3"],r["score3_prob"]),(r["score4"],r["score4_prob"]),
        (r["score5"],r["score5_prob"]),
    ]):
        print(f"    #{i+1}  {s}  ({sp}%)")
    if r["h2h_matches"] > 0:
        print(f"  {SEP}")
        print(f"  歷史交手：共 {r['h2h_matches']} 場，{r['home']} 勝率 {r['h2h_home_wr']}%")
    print(f"  {SEP}")


# ════════════════════════════════════════════════════════════════
# Excel 輸出
# ════════════════════════════════════════════════════════════════

def save_to_excel(results):
    """將所有預測輸出到格式化的 Excel"""
    Path(OUTPUT_EXCEL).parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for r in results:
        rows.append({
            "時間":       r["timestamp"],
            "主隊":       r["home"],
            "客隊":       r["away"],
            "主隊 ELO":   r["home_elo"],
            "客隊 ELO":   r["away_elo"],
            "ELO 差":     round(r["home_elo"] - r["away_elo"], 0),
            "主場勝%":    r["p_home"],
            "平局%":      r["p_draw"],
            "客場勝%":    r["p_away"],
            "預測結果":   r["pred_result"],
            "預測勝隊":   r["winner"],
            "#1 比分":    r["score1"],
            "#1 機率%":   r["score1_prob"],
            "#2 比分":    r["score2"],
            "#2 機率%":   r["score2_prob"],
            "#3 比分":    r["score3"],
            "#3 機率%":   r["score3_prob"],
            "#4 比分":    r["score4"],
            "#4 機率%":   r["score4_prob"],
            "#5 比分":    r["score5"],
            "#5 機率%":   r["score5_prob"],
            "歷史交手場次": r["h2h_matches"],
            "主隊歷史勝率%": r["h2h_home_wr"],
            "λ 主場":     r["lambda_home"],
            "λ 客場":     r["lambda_away"],
        })

    df = pd.DataFrame(rows)

    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        # 若已有檔案，讀入已有資料合併
        if Path(OUTPUT_EXCEL).exists():
            old_df = pd.read_excel(OUTPUT_EXCEL)
            df = pd.concat([old_df, df], ignore_index=True)

        with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="預測結果")
            ws = writer.sheets["預測結果"]

            # 欄寬
            col_widths = {
                "A":16, "B":20, "C":20, "D":10, "E":10, "F":8,
                "G":9,  "H":8,  "I":9,  "J":10, "K":16,
                "L":9,  "M":9,  "N":9,  "O":9,  "P":9,  "Q":9,
                "R":9,  "S":9,  "T":9,  "U":9,
                "V":12, "W":14, "X":8,  "Y":8,
            }
            for col, w in col_widths.items():
                ws.column_dimensions[col].width = w

            # 標題列樣式
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin   = Side(style="thin", color="BFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center
                cell.border    = border

            ws.row_dimensions[1].height = 32

            # 資料列樣式
            fill_even = PatternFill("solid", fgColor="F2F7FF")
            fill_home = PatternFill("solid", fgColor="E8F5E9")  # 主場勝 → 綠
            fill_away = PatternFill("solid", fgColor="FFF3E0")  # 客場勝 → 橙
            fill_draw = PatternFill("solid", fgColor="F3F3F3")  # 平局  → 灰

            result_col_idx = df.columns.get_loc("預測結果") + 1

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                result_val = str(ws.cell(row=row_idx, column=result_col_idx).value or "")
                if   "主場勝" in result_val: row_fill = fill_home
                elif "客場勝" in result_val: row_fill = fill_away
                elif "平局"   in result_val: row_fill = fill_draw
                else: row_fill = fill_even if row_idx % 2 == 0 else None

                for cell in row:
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font      = Font(size=10, name="Arial")
                    if row_fill:
                        cell.fill = row_fill

            # 凍結首列
            ws.freeze_panes = "A2"

        print(f"\n✅ 已輸出至 {OUTPUT_EXCEL}（共 {len(df)} 筆）")

    except ImportError:
        # fallback：沒有 openpyxl 就存 CSV
        csv_path = OUTPUT_EXCEL.replace(".xlsx", ".csv")
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"\n⚠ 未安裝 openpyxl，已改存為 CSV：{csv_path}")
        print(f"  安裝方式：pip install openpyxl")


# ════════════════════════════════════════════════════════════════
# 主程式
# ════════════════════════════════════════════════════════════════

def main():
    print("\n" + "="*55)
    print("  2026 World Cup 對戰預測工具")
    print("  輸入 ? 列出所有隊伍，輸入 q 結束並儲存")
    print("="*55)

    xgb_data, poi_data, ens_data, feat, elo, h2h_dict, teams = load_all()

    results = []

    while True:
        print(f"\n  目前已預測 {len(results)} 場對陣")
        print("  ─────────────────────────────────")

        # 選主隊
        home_input = input("\n  選擇主隊（Enter 直接選隊，q 儲存並結束）：").strip()
        if home_input.lower() == "q":
            break
        if home_input == "?":
            print_team_list(teams)
            home_input = input("\n  選擇主隊（輸入編號或隊名）：").strip()

        # 解析主隊
        if home_input.isdigit():
            idx = int(home_input) - 1
            if not (0 <= idx < len(teams)):
                print(f"  ⚠ 請輸入 1~{len(teams)} 之間的編號"); continue
            home = teams[idx]
        else:
            res = fuzzy_find_team(home_input, teams)
            if res is None:
                print(f"  ⚠ 找不到「{home_input}」"); continue
            if isinstance(res, list):
                print(f"  找到多個：{', '.join(res)}")
                sub = input("  請輸入編號選擇：").strip()
                if sub.isdigit() and 1 <= int(sub) <= len(res):
                    home = res[int(sub)-1]
                else:
                    continue
            else:
                home = res
        print(f"  ✅ 主隊：{home}")

        # 選客隊
        away_input = input(f"\n  選擇客隊（vs {home}）：").strip()
        if away_input.lower() == "q":
            break
        if away_input == "?":
            print_team_list(teams)
            away_input = input("\n  選擇客隊（輸入編號或隊名）：").strip()

        if away_input.isdigit():
            idx = int(away_input) - 1
            if not (0 <= idx < len(teams)):
                print(f"  ⚠ 請輸入 1~{len(teams)} 之間的編號"); continue
            away = teams[idx]
        else:
            res = fuzzy_find_team(away_input, teams)
            if res is None:
                print(f"  ⚠ 找不到「{away_input}」"); continue
            if isinstance(res, list):
                print(f"  找到多個：{', '.join(res)}")
                sub = input("  請輸入編號選擇：").strip()
                if sub.isdigit() and 1 <= int(sub) <= len(res):
                    away = res[int(sub)-1]
                else:
                    continue
            else:
                away = res

        if away == home:
            print("  ⚠ 主客隊不能相同"); continue
        print(f"  ✅ 客隊：{away}")

        # 執行預測
        print("\n  🔍 預測中...")
        r = predict(home, away, xgb_data, poi_data, ens_data, feat, elo, h2h_dict)
        print_result(r)
        results.append(r)

        # 是否繼續
        again = input("\n  繼續預測下一場？（Enter 繼續 / q 儲存結束）：").strip().lower()
        if again == "q":
            break

    # 儲存
    if results:
        save_to_excel(results)
    else:
        print("\n  沒有預測記錄，結束。")

    print("\n  再見！\n")


if __name__ == "__main__":
    main()
